#!/usr/bin/env python3
"""
Creates a weekly exhibit Excel file showing the most recent 5 weeks of NY gaming data.
Layout: Week | Handle / GGR / Hold per operator (sorted by handle desc) + Statewide
Each week has two rows: actual data row + yy increase row with signed % / bps.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ── Colours ──────────────────────────────────────────────────────────────────
NAVY        = "1F3864"
LIGHT_GRAY  = "D6DCE4"
GREEN       = "00B050"
RED         = "FF0000"
BLACK       = "000000"
GRAY_TEXT   = "595959"
WHITE       = "FFFFFF"

HEADER_FILL     = PatternFill(start_color=NAVY,       end_color=NAVY,       fill_type="solid")
YOY_FILL        = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
WHITE_FILL      = PatternFill(start_color=WHITE,      end_color=WHITE,      fill_type="solid")

HEADER_FONT     = Font(color=WHITE,     bold=True,  size=10, name="Arial")
DATA_FONT       = Font(color=BLACK,                 size=10, name="Arial")
BOLD_DATA_FONT  = Font(color=BLACK,     bold=True,  size=10, name="Arial")
POS_FONT        = Font(color=GREEN,                 size=9,  name="Arial")
NEG_FONT        = Font(color=RED,                   size=9,  name="Arial")
YOY_LABEL_FONT  = Font(color=GRAY_TEXT, italic=True, size=9, name="Arial")

# ── Brand display name shortening ─────────────────────────────────────────────
BRAND_DISPLAY = {
    'Bally Bet':                 'Bally Bet',
    'BetMGM':                    'BetMGM',
    'Caesars Sport Book':        'Caesars',
    'DraftKings Sport Book':     'DraftKings',
    'ESPN Bet':                  'ESPN Bet',
    'Wynn Interactive':          'Wynn',
    'FanDuel':                   'FanDuel',
    'Fanatics':                  'Fanatics',
    'Resorts World Bet':         'Resorts World',
    'Rush Street Interactive':   'Rush Street',
}

def shorten(brand: str) -> str:
    return BRAND_DISPLAY.get(brand, brand)

# ── Border helpers ────────────────────────────────────────────────────────────
def _cell_border(left=None, right=None, top=None, bottom=None):
    def s(c):
        return Side(style="thin", color=c) if c else Side(style=None)
    return Border(left=s(left), right=s(right), top=s(top), bottom=s(bottom))

INNER_BORDER  = _cell_border("BFBFBF", "BFBFBF", "BFBFBF", "BFBFBF")
HEADER_BORDER = _cell_border("FFFFFF", "FFFFFF", "FFFFFF", "FFFFFF")

def apply_borders(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = INNER_BORDER

# ── YoY date lookup ──────────────────────────────────────────────────────────
def find_yoy_date(current_date, all_dates):
    """Closest date ~364 days ago within a ±7-day window."""
    target = current_date - pd.DateOffset(days=364)
    window = [d for d in all_dates if abs((d - target).days) <= 7]
    if not window:
        return None
    return min(window, key=lambda x: abs((x - target).days))

# ── Formatting helpers ────────────────────────────────────────────────────────
def fmt_yoy_pct(val) -> str | None:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    sign = '+' if val > 0 else ''
    return f"{sign}{val * 100:.0f}%"

def fmt_yoy_bps(diff) -> str | None:
    """diff is raw decimal change in hold (e.g. 0.0426 → +426bps)."""
    if diff is None or (isinstance(diff, float) and np.isnan(diff)):
        return None
    bps = diff * 10000
    sign = '+' if bps > 0 else ''
    return f"{sign}{bps:.0f}bps"

# ── Main function ─────────────────────────────────────────────────────────────
def create_weekly_exhibit(data_file: str = 'ny_gaming_data.csv',
                          output_file: str = 'ny_gaming_weekly_exhibit.xlsx',
                          num_weeks: int = 5) -> str:
    logger.info("📋 Creating weekly exhibit...")

    df = pd.read_csv(data_file)
    df['Date'] = pd.to_datetime(df['Date'])

    all_dates    = sorted(df['Date'].unique(), reverse=True)
    recent_dates = all_dates[:num_weeks]

    # Build pivot tables
    handle_piv = df.pivot_table(index='Date', columns='Brand', values='Handle', aggfunc='sum')
    ggr_piv    = df.pivot_table(index='Date', columns='Brand', values='GGR',    aggfunc='sum')

    handle_piv['Statewide'] = handle_piv.sum(axis=1)
    ggr_piv['Statewide']    = ggr_piv.sum(axis=1)
    hold_piv = ggr_piv / handle_piv.replace(0, np.nan)

    # Only show these four operators (matching the screenshot) + Statewide
    FEATURED_BRANDS = [
        'DraftKings Sport Book',
        'FanDuel',
        'BetMGM',
        'Fanatics',
    ]
    available = set(df['Brand'].unique())
    brands   = [b for b in FEATURED_BRANDS if b in available]
    all_cols = brands + ['Statewide']

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Weekly Exhibit'

    num_groups = len(all_cols)

    # ── Row 1: brand / group headers ──────────────────────────────────────────
    c = ws.cell(row=1, column=1, value="Week")
    c.fill      = HEADER_FILL
    c.font      = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    # Merge Week header over 2 rows
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    for idx, col_name in enumerate(all_cols):
        sc    = 2 + idx * 3
        label = 'Statewide' if col_name == 'Statewide' else shorten(col_name)
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc + 2)
        c = ws.cell(row=1, column=sc, value=label)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')

    # ── Row 2: Handle / GGR / Hold sub-headers ────────────────────────────────
    # Week cell already merged; fill it
    ws.cell(row=2, column=1).fill      = HEADER_FILL
    ws.cell(row=2, column=1).font      = HEADER_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')

    for idx in range(num_groups):
        for j, metric in enumerate(['Handle', 'GGR', 'Hold']):
            col = 2 + idx * 3 + j
            c = ws.cell(row=2, column=col, value=metric)
            c.fill      = HEADER_FILL
            c.font      = HEADER_FONT
            c.alignment = Alignment(horizontal='center', vertical='center')

    # ── Data + YoY rows ───────────────────────────────────────────────────────
    cur_row = 3

    def safe_get(piv, date, col):
        try:
            if date in piv.index and col in piv.columns:
                v = piv.at[date, col]
                return v if not (isinstance(v, float) and np.isnan(v)) else None
        except Exception:
            pass
        return None

    for date in recent_dates:
        date_str = f"{date.month}/{date.day}/{date.year}"
        yoy_date = find_yoy_date(date, all_dates)

        # ── Data row ──────────────────────────────────────────────────────────
        d = ws.cell(row=cur_row, column=1, value=date_str)
        d.fill      = WHITE_FILL
        d.font      = BOLD_DATA_FONT
        d.alignment = Alignment(horizontal='center', vertical='center')

        for idx, col_name in enumerate(all_cols):
            sc = 2 + idx * 3
            hv = safe_get(handle_piv, date, col_name)
            gv = safe_get(ggr_piv,    date, col_name)
            ho = safe_get(hold_piv,   date, col_name)

            h_cell = ws.cell(row=cur_row, column=sc)
            h_cell.value         = int(round(hv)) if hv is not None else None
            h_cell.number_format = '#,##0'
            h_cell.font          = DATA_FONT
            h_cell.fill          = WHITE_FILL
            h_cell.alignment     = Alignment(horizontal='center', vertical='center')

            g_cell = ws.cell(row=cur_row, column=sc + 1)
            g_cell.value         = int(round(gv)) if gv is not None else None
            g_cell.number_format = '#,##0'
            g_cell.font          = DATA_FONT
            g_cell.fill          = WHITE_FILL
            g_cell.alignment     = Alignment(horizontal='center', vertical='center')

            hold_cell = ws.cell(row=cur_row, column=sc + 2)
            if ho is not None:
                hold_cell.value         = ho
                hold_cell.number_format = '0.0%'
            hold_cell.font      = DATA_FONT
            hold_cell.fill      = WHITE_FILL
            hold_cell.alignment = Alignment(horizontal='center', vertical='center')

        cur_row += 1

        # ── YoY row ───────────────────────────────────────────────────────────
        yl = ws.cell(row=cur_row, column=1, value="yy increase")
        yl.fill      = YOY_FILL
        yl.font      = YOY_LABEL_FONT
        yl.alignment = Alignment(horizontal='center', vertical='center')

        for idx, col_name in enumerate(all_cols):
            sc = 2 + idx * 3
            for j in range(3):
                ws.cell(row=cur_row, column=sc + j).fill = YOY_FILL

            if yoy_date is not None:
                ch = safe_get(handle_piv, date,     col_name)
                ph = safe_get(handle_piv, yoy_date, col_name)
                cg = safe_get(ggr_piv,    date,     col_name)
                pg = safe_get(ggr_piv,    yoy_date, col_name)
                co = safe_get(hold_piv,   date,     col_name)
                po = safe_get(hold_piv,   yoy_date, col_name)

                handle_yoy = (ch - ph) / ph if (ch and ph and ph != 0) else None
                ggr_yoy    = (cg - pg) / pg if (cg and pg and pg != 0) else None
                hold_diff  = (co - po)      if (co is not None and po is not None) else None

                # Handle YoY
                h_yoy = ws.cell(row=cur_row, column=sc)
                txt = fmt_yoy_pct(handle_yoy)
                if txt:
                    h_yoy.value = txt
                    h_yoy.font  = POS_FONT if handle_yoy >= 0 else NEG_FONT
                h_yoy.alignment = Alignment(horizontal='center', vertical='center')

                # GGR YoY
                g_yoy = ws.cell(row=cur_row, column=sc + 1)
                txt = fmt_yoy_pct(ggr_yoy)
                if txt:
                    g_yoy.value = txt
                    g_yoy.font  = POS_FONT if ggr_yoy >= 0 else NEG_FONT
                g_yoy.alignment = Alignment(horizontal='center', vertical='center')

                # Hold YoY bps
                ho_yoy = ws.cell(row=cur_row, column=sc + 2)
                txt = fmt_yoy_bps(hold_diff)
                if txt:
                    ho_yoy.value = txt
                    ho_yoy.font  = POS_FONT if hold_diff >= 0 else NEG_FONT
                ho_yoy.alignment = Alignment(horizontal='center', vertical='center')

        cur_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 13
    for idx in range(num_groups):
        for j, metric in enumerate(['Handle', 'GGR', 'Hold']):
            col_letter = get_column_letter(2 + idx * 3 + j)
            ws.column_dimensions[col_letter].width = 15 if metric in ('Handle', 'GGR') else 9

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for r in range(3, cur_row):
        ws.row_dimensions[r].height = 16

    # ── Borders ───────────────────────────────────────────────────────────────
    last_col = 1 + num_groups * 3
    apply_borders(ws, 1, cur_row - 1, 1, last_col)

    # ── Freeze header rows ────────────────────────────────────────────────────
    ws.freeze_panes = 'B3'

    wb.save(output_file)
    logger.info(f"✅ Weekly exhibit saved: {output_file}")
    return output_file


if __name__ == "__main__":
    create_weekly_exhibit()
